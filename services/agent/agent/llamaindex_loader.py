"""
LlamaIndex Document Loaders — multi-format file parsing.

Supports: PDF, DOCX, XLSX, CSV, HTML, Markdown, PPTX, EPUB, and web URLs.
Uses LlamaIndex's file readers for intelligent parsing that preserves
tables, formatting, and structure (unlike raw text splitting).
"""

import os
import logging
import tempfile
from typing import Optional

logger = logging.getLogger("agent-service.llamaindex-loader")

# Supported file types → loader mappings
SUPPORTED_EXTENSIONS = {
    ".pdf",
    ".docx",
    ".doc",
    ".xlsx",
    ".xls",
    ".csv",
    ".html",
    ".htm",
    ".md",
    ".txt",
    ".pptx",
    ".epub",
    ".json",
    ".xml",
}


def parse_file(file_path: str, metadata: Optional[dict] = None) -> list[dict]:
    """Parse a file using LlamaIndex readers. Returns list of {text, metadata}."""
    from llama_index.readers.file import (
        PDFReader,
        DocxReader,
        CSVReader,
        HTMLTagReader,
        MarkdownReader,
        PptxReader,
        EpubReader,
    )

    ext = os.path.splitext(file_path)[1].lower()
    extra_meta = metadata or {}

    readers = {
        ".pdf": PDFReader,
        ".docx": DocxReader,
        ".doc": DocxReader,
        ".csv": CSVReader,
        ".html": HTMLTagReader,
        ".htm": HTMLTagReader,
        ".md": MarkdownReader,
        ".pptx": PptxReader,
        ".epub": EpubReader,
    }

    documents = []

    if ext in readers:
        reader = readers[ext]()
        try:
            docs = reader.load_data(file_path)
            for doc in docs:
                doc_meta = {
                    **extra_meta,
                    "source": os.path.basename(file_path),
                    "file_type": ext,
                }
                doc_meta.update(doc.metadata or {})
                documents.append({"text": doc.text, "metadata": doc_meta})
        except Exception as e:
            logger.error("LlamaIndex reader failed for %s: %s", file_path, e)
            # Fallback to plain text
            documents = _fallback_read(file_path, extra_meta, ext)
    elif ext in (".xlsx", ".xls"):
        documents = _parse_spreadsheet(file_path, extra_meta, ext)
    elif ext in (".json", ".xml"):
        documents = _fallback_read(file_path, extra_meta, ext)
    elif ext == ".txt" or ext == "":
        documents = _fallback_read(file_path, extra_meta, ext or ".txt")
    else:
        documents = _fallback_read(file_path, extra_meta, ext)

    logger.info("Parsed %s → %d document sections", file_path, len(documents))
    return documents


def _parse_spreadsheet(file_path: str, extra_meta: dict, ext: str) -> list[dict]:
    """Parse Excel files sheet by sheet."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        documents = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_str = "\t".join(str(c) if c is not None else "" for c in row)
                rows.append(row_str)
            text = "\n".join(rows)
            if text.strip():
                meta = {
                    **extra_meta,
                    "source": os.path.basename(file_path),
                    "file_type": ext,
                    "sheet": sheet_name,
                }
                documents.append({"text": text, "metadata": meta})
        wb.close()
        return documents
    except ImportError:
        logger.warning(
            "openpyxl not installed, falling back to plain read for %s", file_path
        )
        return _fallback_read(file_path, extra_meta, ext)
    except Exception as e:
        logger.error("Spreadsheet parse failed for %s: %s", file_path, e)
        return _fallback_read(file_path, extra_meta, ext)


def _fallback_read(file_path: str, extra_meta: dict, ext: str) -> list[dict]:
    """Plain text fallback reader."""
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            text = f.read()
        return [
            {
                "text": text,
                "metadata": {
                    **extra_meta,
                    "source": os.path.basename(file_path),
                    "file_type": ext,
                },
            }
        ]
    except Exception as e:
        logger.error("Fallback read failed for %s: %s", file_path, e)
        return []


def parse_file_bytes(
    content: bytes, filename: str, metadata: Optional[dict] = None
) -> list[dict]:
    """Parse file content from bytes (e.g. uploaded files)."""
    with tempfile.NamedTemporaryFile(
        delete=False, suffix=os.path.splitext(filename)[1]
    ) as tmp:
        tmp.write(content)
        tmp_path = tmp.name
    try:
        return parse_file(
            tmp_path, metadata={**(metadata or {}), "original_filename": filename}
        )
    finally:
        os.unlink(tmp_path)


def parse_url(url: str, metadata: Optional[dict] = None) -> list[dict]:
    """Parse web page content using LlamaIndex web reader."""
    from llama_index.readers.web import SimpleWebPageReader

    try:
        reader = SimpleWebPageReader(html_to_text=True)
        docs = reader.load_data([url])
        extra_meta = metadata or {}
        documents = []
        for doc in docs:
            doc_meta = {**extra_meta, "source": url, "file_type": "web"}
            doc_meta.update(doc.metadata or {})
            documents.append({"text": doc.text, "metadata": doc_meta})
        logger.info("Parsed URL %s → %d sections", url, len(documents))
        return documents
    except Exception as e:
        logger.error("Web parse failed for %s: %s", url, e)
        return [{"text": "", "metadata": {"source": url, "error": str(e)}}]


def get_supported_types() -> list[str]:
    """Return list of supported file extensions."""
    return sorted(SUPPORTED_EXTENSIONS)
